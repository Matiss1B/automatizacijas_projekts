from linkedList import LinkedList
import os
from openpyxl import Workbook, load_workbook
from datetime import datetime

def update_excel(emails):
    file_path = 'excel/mails.xlsx'

    os.makedirs(os.path.dirname(file_path), exist_ok=True)

    if os.path.exists(file_path):
        wb = load_workbook(file_path)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws['A1'] = 'Email'
        ws['B1'] = 'Date Sent'

    existing_emails = {
        (ws[f'A{i}'].value or '').strip().lower()
        for i in range(2, ws.max_row + 1)
    }

    row = ws.max_row + 1
    for email in emails:
        normalized_email = email.strip().lower()
        if normalized_email not in existing_emails:
            ws[f'A{row}'] = email.strip()
            ws[f'B{row}'] = None
            row += 1

    wb.save(file_path)
    wb.close()
    print('Excel updated.')

def create_email_linked_list(template_name):
    wb = load_workbook('excel/'+template_name)
    ws = wb.active

    email_list = LinkedList()

    for row in range(2, ws.max_row + 1):
        email = ws[f'A{row}'].value
        date_sent = ws[f'B{row}'].value
        if email and not date_sent:
            email_list.append(email)

    wb.close()
    return email_list
def mark_emails_as_sent(email_list, template_name='mails.xlsx'):
    wb = load_workbook('excel/' + template_name)
    ws = wb.active
    current_date = datetime.now().strftime("%Y-%m-%d")

    for row in range(2, ws.max_row + 1):
        email = ws[f'A{row}'].value
        if email in email_list and not ws[f'B{row}'].value:
            ws[f'B{row}'] = current_date

    wb.save('excel/' + template_name)
    wb.close()
    print("Emails updated with send date.")